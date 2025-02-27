from django.shortcuts import render
from django.http import HttpResponse
import openpyxl
from register.models import Participant, Bill, ExpoDetail, CyberShowDetail, FragFestDetail
from .models import Event
import json


from django.views.generic import TemplateView


def export_participants_to_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Participants"

    # Define the headers
    headers = ['ID', 'Name', 'Phone', 'Email', 'College Name', 'College ID', 'Course', 'Event', 'Amount', 'Paid', 'Payment ID', 'Time']
    ws.append(headers)

    # Fetch participant data from the database
    participants = Participant.objects.all()
    
    # Populate the worksheet with participant data
    for participant in participants:
        try:
            participants_bill = Bill.objects.get(participant=participant)
        except Bill.DoesNotExist:
            participants_bill = None
        try:
            participant_event = Event.objects.get(name=participant.event)
        except Event.DoesNotExist:
            participant_event = None

        ws.append([
            participant.id,
            participant.name,
            participant.phone,
            participant.email,
            participant.college_name,
            participant.college_id,
            participant.course,
            participant.event,
            participant_event.fees if participant_event else None,
            True if participants_bill else False,
            participants_bill.razorpay_order_id if participants_bill else None,
            participant.registered_at.strftime('%Y-%m-%d, %H:%M:%S'),
        ])

    # Add Expo Details sheet
    expodetails = wb.create_sheet(title="Expo Details")
    expo_headers = ['ID', 'Team Name', 'Number of Members', 'Project Detail', 'Project Detail File']
    expodetails.append(expo_headers)
    expo_details = ExpoDetail.objects.all()
    for expo_detail in expo_details:
        expodetails.append([
            expo_detail.participant.id,
            expo_detail.team_name,
            expo_detail.number_of_members,
            expo_detail.project_detail,
            expo_detail.project_detail_file.url,
        ])

    # Add CyberShow Details sheet
    cybershowdetails = wb.create_sheet(title="CyberShow Details")
    cybershow_headers = ['ID', 'Team Name', 'Number of Members', 'Skit Detail File', 'Participant Names']
    cybershowdetails.append(cybershow_headers)
    cybershow_details = CyberShowDetail.objects.all()
    for cybershow_detail in cybershow_details:
        par_names = ""
        for members in cybershow_detail.participant_names:
            par_names += members + ", "
        cybershowdetails.append([
                cybershow_detail.participant.id,
                cybershow_detail.team_name,
                cybershow_detail.number_of_members,
                cybershow_detail.skit_detail_file.url,
                par_names,
            ])
            
        
    
    fragfestdetails = wb.create_sheet(title="FragFest Details")
    fragfest_headers = ['ID', 'Members']
    fragfestdetails.append(fragfest_headers)

    fragfest_details = FragFestDetail.objects.all()
    for fragfest_detail in fragfest_details:
        mem_names = ""
        for member in fragfest_detail.team_member_details:
            mem_names += f'username: {member.get("user_name")}, real name: {member.get("real_name")}, '
        fragfestdetails.append([
            fragfest_detail.participant.id,
            mem_names
            ])

    # Set the response content type and headers
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=participants.xlsx'

    # Save the workbook to the response
    wb.save(response)

    return response



class TermsAndConditionsView(TemplateView):
    template_name = 'terms_and_conditions.html'


class PrivacyPolicyView(TemplateView):
    template_name = 'privacy_policy.html'

class RefundPolicyView(TemplateView):
    template_name = 'refund_policy.html'
